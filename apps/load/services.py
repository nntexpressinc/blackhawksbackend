import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.core.files.base import ContentFile
import io
from collections import defaultdict
from apps.load.models.ifta import StateTaxRate


class IFTAReportProcessor:
    def __init__(self, ifta_report_instance):
        self.ifta_report = ifta_report_instance
        self.fuel_data = None
        self.mile_data = None
        self.unit_mpg = {}
        self.overall_mpg = 0
        
    def process_files(self):
        """Main method to process both files and generate the result Excel file"""
        try:
            # Process fuel Excel file
            self.fuel_data = self._process_fuel_file()
            
            # Process mile CSV file
            self.mile_data = self._process_mile_file()
            
            # Calculate MPG values
            self._calculate_mpg()
            
            # Generate result Excel file
            result_file = self._generate_result_excel()
            
            # Save result file to model
            self.ifta_report.result_file = result_file
            self.ifta_report.save()
            
            return True
            
        except Exception as e:
            print(f"Error processing files: {str(e)}")
            raise e
    
    def _process_fuel_file(self):
        """Process the fuel Excel file and extract required data"""
        # Read Excel file
        fuel_df = pd.read_excel(self.ifta_report.fuel.path)
        
        # Extract required columns
        required_columns = ['Unit', 'Quantity', 'LocationState']
        
        # Check if required columns exist
        for col in required_columns:
            if col not in fuel_df.columns:
                raise ValueError(f"Required column '{col}' not found in fuel Excel file")
        
        # Clean data - remove any null values
        fuel_df = fuel_df.dropna(subset=required_columns)
        
        # Group by LocationState and sum Quantity (GALLON PER STATE)
        state_quantity = fuel_df.groupby('LocationState')['Quantity'].sum().to_dict()
        
        # Group by Unit and LocationState, sum Quantity
        unit_state_quantity = fuel_df.groupby(['Unit', 'LocationState'])['Quantity'].sum().reset_index()
        
        # Group by Unit only, sum Quantity
        unit_total_quantity = fuel_df.groupby('Unit')['Quantity'].sum().to_dict()
        
        # Get unique units
        unique_units = sorted(fuel_df['Unit'].unique())
        
        return {
            'state_quantity': state_quantity,
            'unit_state_quantity': unit_state_quantity,
            'unit_total_quantity': unit_total_quantity,
            'unique_units': unique_units,
            'total_quantity': fuel_df['Quantity'].sum()
        }
    
    def _process_mile_file(self):
        """Process the mile CSV file with new format (Unit, State, Miles)"""
        # Read CSV file
        mile_df = pd.read_csv(self.ifta_report.mile.path)
        
        # Extract required columns
        required_columns = ['Unit', 'State', 'Miles']
        
        # Check if required columns exist
        for col in required_columns:
            if col not in mile_df.columns:
                raise ValueError(f"Required column '{col}' not found in mile CSV file")
        
        # Clean data - remove any null values
        mile_df = mile_df.dropna(subset=required_columns)
        
        # Group by State and sum Miles (MILEAGE PER STATE)
        state_miles = mile_df.groupby('State')['Miles'].sum().to_dict()
        
        # Group by Unit and State, sum Miles
        unit_state_miles = mile_df.groupby(['Unit', 'State'])['Miles'].sum().reset_index()
        
        # Group by Unit only, sum Miles
        unit_total_miles = mile_df.groupby('Unit')['Miles'].sum().to_dict()
        
        return {
            'state_miles': state_miles,
            'unit_state_miles': unit_state_miles,
            'unit_total_miles': unit_total_miles,
            'total_miles': mile_df['Miles'].sum()
        }
    
    def _calculate_mpg(self):
        """Calculate MPG for each unit and overall MPG"""
        # Calculate MPG for each unit
        for unit in self.fuel_data['unique_units']:
            unit_miles = self.mile_data['unit_total_miles'].get(unit, 0)
            unit_quantity = self.fuel_data['unit_total_quantity'].get(unit, 0)
            
            if unit_quantity > 0:
                self.unit_mpg[unit] = unit_miles / unit_quantity
            else:
                self.unit_mpg[unit] = 0
        
        # Calculate overall MPG
        if self.fuel_data['total_quantity'] > 0:
            self.overall_mpg = self.mile_data['total_miles'] / self.fuel_data['total_quantity']
        else:
            self.overall_mpg = 0
    
    def _get_tax_rate(self, state):
        """Get tax rate for a specific state and quarter"""
        try:
            tax_rate_obj = StateTaxRate.objects.get(
                state=state, 
                quorter=self.ifta_report.quorter
            )
            return float(tax_rate_obj.tax_rate)
        except StateTaxRate.DoesNotExist:
            print(f"Tax rate not found for state {state} and quarter {self.ifta_report.quorter}")
            return 0.0
    
    def _generate_result_excel(self):
        """Generate the result Excel file with the new table structure"""
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"IFTA Report - {self.ifta_report.quorter}"
        
        # Set up styles
        header_font = Font(bold=True, size=11)
        total_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        current_row = 1
        
        # Table 1: Summary Table per State and Unit
        current_row = self._add_summary_table(ws, current_row, header_font, total_font, border, green_fill)
        current_row += 2
        
        # Table 2: Per-State Tax Table (next to Table 1, but we'll put it below for simplicity)
        current_row = self._add_state_tax_table(ws, current_row, header_font, total_font, border)
        current_row += 2
        
        # Table 3: Detailed Tax Table per Unit
        current_row = self._add_detailed_tax_table(ws, current_row, header_font, total_font, border)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create Django file
        filename = f"ifta_report_{self.ifta_report.quorter.replace(' ', '_').lower()}.xlsx"
        return ContentFile(excel_buffer.getvalue(), name=filename)
    
    def _add_summary_table(self, ws, start_row, header_font, total_font, border, green_fill):
        """Add Table 1: Summary Table per State and Unit - Use decimals except for GALLON PER STATE and MILEAGE PER STATE"""
        # Title
        ws.cell(row=start_row, column=1, value="Table 1: Summary per State and Unit")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=14)
        start_row += 2
        
        # Get all states from both fuel and mile data
        all_states = set(self.fuel_data['state_quantity'].keys()) | set(self.mile_data['state_miles'].keys())
        all_states = sorted(all_states)
        unique_units = self.fuel_data['unique_units']
        
        # Headers
        col = 1
        ws.cell(row=start_row, column=col, value="State")
        ws.cell(row=start_row, column=col).font = header_font
        ws.cell(row=start_row, column=col).border = border
        col += 1
        
        # Unit columns (quantity and miles for each unit)
        for unit in unique_units:
            # Unit quantity column
            ws.cell(row=start_row, column=col, value=str(unit))
            ws.cell(row=start_row, column=col).font = header_font
            ws.cell(row=start_row, column=col).border = border
            col += 1
            
            # Unit mileage column
            ws.cell(row=start_row, column=col, value=f"{unit} mileage")
            ws.cell(row=start_row, column=col).font = header_font
            ws.cell(row=start_row, column=col).border = border
            col += 1
        
        # GALLON PER STATE and MILEAGE PER STATE columns
        ws.cell(row=start_row, column=col, value="GALLON PER STATE")
        ws.cell(row=start_row, column=col).font = header_font
        ws.cell(row=start_row, column=col).border = border
        ws.cell(row=start_row, column=col).fill = green_fill
        col += 1
        
        ws.cell(row=start_row, column=col, value="MILEAGE PER STATE")
        ws.cell(row=start_row, column=col).font = header_font
        ws.cell(row=start_row, column=col).border = border
        ws.cell(row=start_row, column=col).fill = green_fill
        
        start_row += 1
        
        # Data rows
        unit_quantity_totals = {unit: 0 for unit in unique_units}
        unit_miles_totals = {unit: 0 for unit in unique_units}
        total_gallon_per_state = 0
        total_mileage_per_state = 0
        
        for state in all_states:
            col = 1
            ws.cell(row=start_row, column=col, value=state)
            ws.cell(row=start_row, column=col).border = border
            col += 1
            
            state_gallon_total = 0
            state_miles_total = 0
            
            for unit in unique_units:
                # Get quantity for this unit and state
                unit_state_df = self.fuel_data['unit_state_quantity']
                quantity = unit_state_df[
                    (unit_state_df['Unit'] == unit) & 
                    (unit_state_df['LocationState'] == state)
                ]['Quantity'].sum()
                
                # Use decimal format (2 decimal places) for unit quantities
                ws.cell(row=start_row, column=col, value=f"{quantity:.2f}" if quantity > 0 else "0.00")
                ws.cell(row=start_row, column=col).border = border
                unit_quantity_totals[unit] += quantity
                state_gallon_total += quantity
                col += 1
                
                # Get miles for this unit and state
                unit_miles_df = self.mile_data['unit_state_miles']
                miles = unit_miles_df[
                    (unit_miles_df['Unit'] == unit) & 
                    (unit_miles_df['State'] == state)
                ]['Miles'].sum()
                
                # Use decimal format (2 decimal places) for unit miles
                ws.cell(row=start_row, column=col, value=f"{miles:.2f}" if miles > 0 else "0.00")
                ws.cell(row=start_row, column=col).border = border
                unit_miles_totals[unit] += miles
                state_miles_total += miles
                col += 1
            
            # GALLON PER STATE - keep as integer
            ws.cell(row=start_row, column=col, value=int(state_gallon_total))
            ws.cell(row=start_row, column=col).border = border
            ws.cell(row=start_row, column=col).fill = green_fill
            total_gallon_per_state += state_gallon_total
            col += 1
            
            # MILEAGE PER STATE - keep as integer
            ws.cell(row=start_row, column=col, value=int(state_miles_total))
            ws.cell(row=start_row, column=col).border = border
            ws.cell(row=start_row, column=col).fill = green_fill
            total_mileage_per_state += state_miles_total
            
            start_row += 1
        
        # Total row
        col = 1
        ws.cell(row=start_row, column=col, value="TOTAL")
        ws.cell(row=start_row, column=col).font = total_font
        ws.cell(row=start_row, column=col).border = border
        col += 1
        
        for unit in unique_units:
            # Unit quantity total - decimal format
            ws.cell(row=start_row, column=col, value=f"{unit_quantity_totals[unit]:.2f}")
            ws.cell(row=start_row, column=col).font = total_font
            ws.cell(row=start_row, column=col).border = border
            col += 1
            
            # Unit miles total - decimal format
            ws.cell(row=start_row, column=col, value=f"{unit_miles_totals[unit]:.2f}")
            ws.cell(row=start_row, column=col).font = total_font
            ws.cell(row=start_row, column=col).border = border
            col += 1
        
        # Total GALLON PER STATE - keep as integer
        ws.cell(row=start_row, column=col, value=int(total_gallon_per_state))
        ws.cell(row=start_row, column=col).font = total_font
        ws.cell(row=start_row, column=col).border = border
        ws.cell(row=start_row, column=col).fill = green_fill
        col += 1
        
        # Total MILEAGE PER STATE - keep as integer
        ws.cell(row=start_row, column=col, value=int(total_mileage_per_state))
        ws.cell(row=start_row, column=col).font = total_font
        ws.cell(row=start_row, column=col).border = border
        ws.cell(row=start_row, column=col).fill = green_fill
        
        # Add MPG row
        start_row += 1
        ws.cell(row=start_row, column=1, value="MPG")
        ws.cell(row=start_row, column=1).font = total_font
        ws.cell(row=start_row, column=1).border = border
        
        # Calculate and display MPG for each unit
        col = 2
        for unit in unique_units:
            # Unit MPG
            mpg = self.unit_mpg.get(unit, 0)
            ws.cell(row=start_row, column=col, value=f"{mpg:.2f}")
            ws.cell(row=start_row, column=col).font = total_font
            ws.cell(row=start_row, column=col).border = border
            col += 2  # Skip miles column
        
        # Overall MPG in the last column
        ws.cell(row=start_row, column=col + 1, value=f"{self.overall_mpg:.2f}")
        ws.cell(row=start_row, column=col + 1).font = total_font
        ws.cell(row=start_row, column=col + 1).border = border
        ws.cell(row=start_row, column=col + 1).fill = green_fill
        
        return start_row + 1
    
    def _add_state_tax_table(self, ws, start_row, header_font, total_font, border):
        """Add Table 2: Per-State Tax Table - Keep all as integers"""
        # Title
        ws.cell(row=start_row, column=1, value="Table 2: Per-State Tax Calculations")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=14)
        start_row += 2
        
        # Headers
        headers = ["State", "Taxible Gallon", "Net Taxible Gallon", "TAX"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=header)
            ws.cell(row=start_row, column=col).font = header_font
            ws.cell(row=start_row, column=col).border = border
        
        start_row += 1
        
        # Get all states
        all_states = set(self.fuel_data['state_quantity'].keys()) | set(self.mile_data['state_miles'].keys())
        all_states = sorted(all_states)
        
        total_taxible = 0
        total_net_taxible = 0
        total_tax = 0
        
        for state in all_states:
            mileage_per_state = self.mile_data['state_miles'].get(state, 0)
            gallon_per_state = self.fuel_data['state_quantity'].get(state, 0)
            
            # Taxible Gallon = MILEAGE PER STATE / overall MPG
            taxible_gallon = mileage_per_state / self.overall_mpg if self.overall_mpg > 0 else 0
            
            # Net Taxible Gallon = Taxible Gallon - GALLON PER STATE
            net_taxible_gallon = taxible_gallon - gallon_per_state
            
            # TAX = Net Taxible Gallon × tax_rate
            tax_rate = self._get_tax_rate(state)
            tax = net_taxible_gallon * tax_rate
            
            # Add to totals
            total_taxible += taxible_gallon
            total_net_taxible += net_taxible_gallon
            total_tax += tax
            
            # Write data - keep as integers as requested
            ws.cell(row=start_row, column=1, value=state)
            ws.cell(row=start_row, column=1).border = border
            
            ws.cell(row=start_row, column=2, value=int(taxible_gallon))
            ws.cell(row=start_row, column=2).border = border
            
            ws.cell(row=start_row, column=3, value=int(net_taxible_gallon))
            ws.cell(row=start_row, column=3).border = border
            
            ws.cell(row=start_row, column=4, value=f"${tax:.2f}")
            ws.cell(row=start_row, column=4).border = border
            
            start_row += 1
        
        # Total row
        ws.cell(row=start_row, column=1, value="TOTAL")
        ws.cell(row=start_row, column=1).font = total_font
        ws.cell(row=start_row, column=1).border = border
        
        ws.cell(row=start_row, column=2, value=int(total_taxible))
        ws.cell(row=start_row, column=2).font = total_font
        ws.cell(row=start_row, column=2).border = border
        
        ws.cell(row=start_row, column=3, value=int(total_net_taxible))
        ws.cell(row=start_row, column=3).font = total_font
        ws.cell(row=start_row, column=3).border = border
        
        ws.cell(row=start_row, column=4, value=f"${total_tax:.2f}")
        ws.cell(row=start_row, column=4).font = total_font
        ws.cell(row=start_row, column=4).border = border
        
        return start_row + 1
    
    def _add_detailed_tax_table(self, ws, start_row, header_font, total_font, border):
        """Add Table 3: Detailed Tax Table per Unit - Use decimal numbers"""
        # Title
        ws.cell(row=start_row, column=1, value="Table 3: Detailed Tax Calculations per Unit")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=14)
        start_row += 2
        
        unique_units = self.fuel_data['unique_units']
        all_states = set(self.fuel_data['state_quantity'].keys()) | set(self.mile_data['state_miles'].keys())
        all_states = sorted(all_states)
        
        # Headers
        col = 1
        ws.cell(row=start_row, column=col, value="State")
        ws.cell(row=start_row, column=col).font = header_font
        ws.cell(row=start_row, column=col).border = border
        col += 1
        
        # For each unit, create 3 columns: Taxible Gallon, Net Taxible Gallon, TAX
        for unit in unique_units:
            ws.cell(row=start_row, column=col, value=f"{unit} Taxible Gallon")
            ws.cell(row=start_row, column=col).font = header_font
            ws.cell(row=start_row, column=col).border = border
            col += 1
            
            ws.cell(row=start_row, column=col, value=f"{unit} Net Taxible Gallon")
            ws.cell(row=start_row, column=col).font = header_font
            ws.cell(row=start_row, column=col).border = border
            col += 1
            
            ws.cell(row=start_row, column=col, value=f"{unit} TAX")
            ws.cell(row=start_row, column=col).font = header_font
            ws.cell(row=start_row, column=col).border = border
            col += 1
        
        start_row += 1
        
        # Initialize totals for each unit
        unit_totals = {unit: {'taxible': 0, 'net_taxible': 0, 'tax': 0} for unit in unique_units}
        
        # Data rows
        for state in all_states:
            col = 1
            ws.cell(row=start_row, column=col, value=state)
            ws.cell(row=start_row, column=col).border = border
            col += 1
            
            tax_rate = self._get_tax_rate(state)
            
            for unit in unique_units:
                # Get miles for this unit and state
                unit_miles_df = self.mile_data['unit_state_miles']
                miles = unit_miles_df[
                    (unit_miles_df['Unit'] == unit) & 
                    (unit_miles_df['State'] == state)
                ]['Miles'].sum()
                
                # Get quantity for this unit and state
                unit_state_df = self.fuel_data['unit_state_quantity']
                quantity = unit_state_df[
                    (unit_state_df['Unit'] == unit) & 
                    (unit_state_df['LocationState'] == state)
                ]['Quantity'].sum()
                
                # Taxible Gallon = Miles for that Unit and State / MPG of that Unit
                unit_mpg = self.unit_mpg.get(unit, 0)
                taxible_gallon = miles / unit_mpg if unit_mpg > 0 else 0
                
                # Net Taxible Gallon = Taxible Gallon - Quantity for that Unit and State
                net_taxible_gallon = taxible_gallon - quantity
                
                # TAX = Net Taxible Gallon × tax_rate
                tax = net_taxible_gallon * tax_rate
                
                # Add to unit totals
                unit_totals[unit]['taxible'] += taxible_gallon
                unit_totals[unit]['net_taxible'] += net_taxible_gallon
                unit_totals[unit]['tax'] += tax
                
                # Write data - use decimal format (2 decimal places)
                ws.cell(row=start_row, column=col, value=f"{taxible_gallon:.2f}")
                ws.cell(row=start_row, column=col).border = border
                col += 1
                
                ws.cell(row=start_row, column=col, value=f"{net_taxible_gallon:.2f}")
                ws.cell(row=start_row, column=col).border = border
                col += 1
                
                ws.cell(row=start_row, column=col, value=f"${tax:.2f}")
                ws.cell(row=start_row, column=col).border = border
                col += 1
            
            start_row += 1
        
        # Total row
        col = 1
        ws.cell(row=start_row, column=col, value="TOTAL")
        ws.cell(row=start_row, column=col).font = total_font
        ws.cell(row=start_row, column=col).border = border
        col += 1
        
        for unit in unique_units:
            # Use decimal format for totals
            ws.cell(row=start_row, column=col, value=f"{unit_totals[unit]['taxible']:.2f}")
            ws.cell(row=start_row, column=col).font = total_font
            ws.cell(row=start_row, column=col).border = border
            col += 1
            
            ws.cell(row=start_row, column=col, value=f"{unit_totals[unit]['net_taxible']:.2f}")
            ws.cell(row=start_row, column=col).font = total_font
            ws.cell(row=start_row, column=col).border = border
            col += 1
            
            ws.cell(row=start_row, column=col, value=f"${unit_totals[unit]['tax']:.2f}")
            ws.cell(row=start_row, column=col).font = total_font
            ws.cell(row=start_row, column=col).border = border
            col += 1
        
        return start_row + 1


def process_ifta_report(ifta_report_instance):
    """
    Function to process IFTA report files
    Call this after both fuel and mile files are uploaded
    """
    processor = IFTAReportProcessor(ifta_report_instance)
    return processor.process_files()